"""Drillhole data validation skill.

The rules are a Python port of the core behavior from the referenced
dh-validation-engine TypeScript app, adapted to SecondBrain workspaces.
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from ..safe_fs import (
    resolve_within_workspace,
    read_text_within_workspace,
    write_text_within_workspace,
)


DEFAULT_CONFIG_PATH = ".secondbrain/dh-validation/config.json"
EMBEDDED_CONFIG_FILENAME = "userConfig.json"

TABLE_TYPES = [
    "COLLAR",
    "SURVEY",
    "LITHOLOGY",
    "ASSAY",
    "MINERALIZATION",
    "OXIDATION",
    "GEOTECH",
    "RQD",
    "VEIN",
    "ALTERATION",
    "DENSITY",
]
INTERVAL_TABLES = [
    "LITHOLOGY",
    "ASSAY",
    "MINERALIZATION",
    "OXIDATION",
    "GEOTECH",
    "RQD",
    "VEIN",
    "ALTERATION",
    "DENSITY",
]


@dataclass(frozen=True)
class ValidationError:
    id: str
    table: str
    row_id: str
    site_id: str
    message: str
    severity: str
    type: str
    column: str | None = None
    file_name: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "table": self.table,
            "rowId": self.row_id,
            "siteId": self.site_id,
            "column": self.column,
            "message": self.message,
            "severity": self.severity,
            "type": self.type,
            "fileName": self.file_name,
        }


@dataclass(frozen=True)
class ValidationSummary:
    total_errors: int
    total_warnings: int
    errors: list[ValidationError]
    report_path: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "totalErrors": self.total_errors,
            "totalWarnings": self.total_warnings,
            "reportPath": self.report_path,
            "errors": [error.to_dict() for error in self.errors],
        }


class DrillholeValidationSkill:
    """Validate drillhole tables from CSV/XLSX files inside a workspace."""

    def __init__(self, workspace: str | Path):
        self.workspace = resolve_within_workspace(workspace, ".")
        self.app_config_path = app_config_path()

    def init_config(self, path: str = DEFAULT_CONFIG_PATH, *, overwrite: bool = False) -> Path:
        return self.app_config_path

    def config_exists(self, path: str = DEFAULT_CONFIG_PATH) -> bool:
        return self.app_config_path.exists()

    def config_payload(self, path: str = DEFAULT_CONFIG_PATH) -> dict[str, Any]:
        return {
            "path": self.app_config_path.as_posix(),
            "scope": "app",
            "exists": self.app_config_path.exists(),
            "config": self.load_config(),
        }

    def load_config(self, path: str | None = None) -> dict[str, Any]:
        if path and path != DEFAULT_CONFIG_PATH:
            raise ValueError("Custom config paths are disabled; embedded app userConfig.json is the source of truth")
        return self._load_app_config()

    def save_config(self, config: dict[str, Any], path: str | None = None) -> Path:
        if path and path != DEFAULT_CONFIG_PATH:
            raise ValueError("Custom config paths are disabled; embedded app userConfig.json is the source of truth")
        validated = validate_config(config)
        content = json.dumps(validated, indent=2, ensure_ascii=False)
        target = self.app_config_path
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(content + "\n", encoding="utf-8")
        return target

    def show_config(self, path: str | None = None) -> str:
        config = self.load_config(path)
        return json.dumps(config, indent=2, ensure_ascii=False)

    def validate(
        self,
        inputs: dict[str, str],
        *,
        config_path: str | None = None,
        out_path: str | None = None,
    ) -> ValidationSummary:
        if config_path:
            raise ValueError("Custom config paths are disabled; embedded app userConfig.json is the source of truth")
        config = self._load_app_config()
        configs = _normalize_configs(config.get("configs", []))
        libraries = config.get("libraries", [])
        data = self._load_tables(inputs)
        table_files = {
            _table_type(table_name): rel_path
            for table_name, rel_path in inputs.items()
            if _table_type(table_name) in TABLE_TYPES
        }

        errors: list[ValidationError] = []
        collars = data.get("COLLAR", [])
        survey = data.get("SURVEY", [])

        for table_type, rows in data.items():
            table_config = configs.get(table_type)
            if table_config:
                errors.extend(_validate_structure(rows, table_config))
                errors.extend(_validate_values(rows, table_config, libraries))

        if collars:
            errors.extend(_validate_collar_depths(collars))

        if collars and survey:
            errors.extend(_validate_integrity(collars, survey, "SURVEY"))
            errors.extend(_validate_survey_eoh(collars, survey))

        for table_type in INTERVAL_TABLES:
            rows = data.get(table_type, [])
            if not rows:
                continue
            if collars:
                errors.extend(_validate_integrity(collars, rows, table_type))
                errors.extend(_validate_eoh(collars, rows, table_type))
            errors.extend(_validate_intervals(rows, table_type))

        errors = [
            replace(error, file_name=table_files.get(error.table, error.table))
            for error in errors
        ]

        summary = ValidationSummary(
            total_errors=sum(1 for error in errors if error.severity == "CRITICAL"),
            total_warnings=sum(1 for error in errors if error.severity == "WARNING"),
            errors=errors,
        )

        if out_path:
            report_path = self._write_report(summary, out_path)
            summary = ValidationSummary(
                total_errors=summary.total_errors,
                total_warnings=summary.total_warnings,
                errors=summary.errors,
                report_path=report_path,
            )

        return summary

    def _load_app_config(self) -> dict[str, Any]:
        data = json.loads(self.app_config_path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            raise ValueError("Embedded drillhole validation config must be a JSON object")
        return data

    def _load_config(self, path: str) -> dict[str, Any]:
        target = resolve_within_workspace(self.workspace, path)
        if not target.exists():
            return default_config()
        raw = read_text_within_workspace(self.workspace, path)
        data = json.loads(raw)
        if not isinstance(data, dict):
            raise ValueError("Drillhole validation config must be a JSON object")
        return data

    def _load_tables(self, inputs: dict[str, str]) -> dict[str, list[dict[str, Any]]]:
        loaded: dict[str, list[dict[str, Any]]] = {}
        for table_name, rel_path in inputs.items():
            table_type = _table_type(table_name)
            if table_type not in TABLE_TYPES:
                raise ValueError(f"Unsupported drillhole table: {table_name}")
            rows = self._read_table(rel_path)
            loaded[table_type] = rows
        return loaded

    def _read_table(self, rel_path: str) -> list[dict[str, Any]]:
        path = resolve_within_workspace(self.workspace, rel_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            text = read_text_within_workspace(self.workspace, rel_path)
            reader = csv.DictReader(io.StringIO(text))
            return _normalize_rows(list(reader))
        if suffix in {".xlsx", ".xlsm"}:
            return _normalize_rows(_read_xlsx_rows(path))
        raise ValueError(f"Unsupported drillhole table file type: {suffix or '(none)'}")

    def _write_report(self, summary: ValidationSummary, out_path: str) -> str:
        suffix = Path(out_path).suffix.lower()
        if suffix == ".json":
            content = json.dumps(summary.to_dict(), indent=2, ensure_ascii=False) + "\n"
        elif suffix in {"", ".md", ".markdown"}:
            content = _render_markdown_report(summary)
        else:
            raise ValueError("Report output must be .json or .md")
        path = write_text_within_workspace(self.workspace, out_path, content)
        return path.relative_to(self.workspace).as_posix()


def default_config() -> dict[str, Any]:
    """Return the embedded app-level drillhole validation config."""
    return _app_default_config()


def app_config_path() -> Path:
    override = os.environ.get("SECONDBRAIN_DH_CONFIG_PATH", "").strip()
    if override:
        return Path(override).expanduser().resolve()
    return Path(__file__).resolve().parent.parent / "defaults" / EMBEDDED_CONFIG_FILENAME


def _app_default_config() -> dict[str, Any]:
    data = json.loads(app_config_path().read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Embedded drillhole validation config must be a JSON object")
    return data


def validate_config(config: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(config, dict):
        raise ValueError("Drillhole validation config must be a JSON object")
    libraries = config.get("libraries", [])
    configs = config.get("configs", [])
    if not isinstance(libraries, list):
        raise ValueError("Drillhole validation config 'libraries' must be an array")
    if not isinstance(configs, list):
        raise ValueError("Drillhole validation config 'configs' must be an array")

    library_ids: set[str] = set()
    for index, library in enumerate(libraries):
        if not isinstance(library, dict):
            raise ValueError(f"Library at index {index} must be an object")
        library_id = str(library.get("id") or "").strip()
        if not library_id:
            raise ValueError(f"Library at index {index} requires an id")
        if library_id in library_ids:
            raise ValueError(f"Duplicate library id: {library_id}")
        library_ids.add(library_id)
        items = library.get("items", [])
        if not isinstance(items, list):
            raise ValueError(f"Library '{library_id}' items must be an array")
        for item_index, item in enumerate(items):
            if not isinstance(item, dict):
                raise ValueError(f"Library '{library_id}' item {item_index} must be an object")
            if not str(item.get("code") or "").strip():
                raise ValueError(f"Library '{library_id}' item {item_index} requires a code")

    for config_index, table_config in enumerate(configs):
        if not isinstance(table_config, dict):
            raise ValueError(f"Table config at index {config_index} must be an object")
        table_type = _table_type(str(table_config.get("tableType") or ""))
        if table_type not in TABLE_TYPES:
            raise ValueError(f"Unknown tableType: {table_config.get('tableType')}")
        table_config["tableType"] = table_type
        columns = table_config.get("columns", [])
        if not isinstance(columns, list):
            raise ValueError(f"Table config '{table_type}' columns must be an array")
        seen_columns: set[str] = set()
        for column_index, column in enumerate(columns):
            if not isinstance(column, dict):
                raise ValueError(f"Column {column_index} in '{table_type}' must be an object")
            column_name = _normalize_column_name(str(column.get("columnName") or ""))
            if not column_name:
                raise ValueError(f"Column {column_index} in '{table_type}' requires columnName")
            if column_name in seen_columns:
                raise ValueError(f"Duplicate column '{column_name}' in '{table_type}'")
            seen_columns.add(column_name)
            column["columnName"] = column_name
            column.setdefault("label", column_name)
            column.setdefault("isSchemaRequired", False)
            column.setdefault("isMandatory", False)
            column.setdefault("type", "string")
            validation = column.get("validation") or {}
            if validation and not isinstance(validation, dict):
                raise ValueError(f"Column '{column_name}' validation must be an object")
            lookup = validation.get("lookup") if isinstance(validation, dict) else None
            if lookup is not None:
                if not isinstance(lookup, dict):
                    raise ValueError(f"Column '{column_name}' lookup validation must be an object")
                library_id = str(lookup.get("libraryId") or "").strip()
                if not library_id:
                    raise ValueError(f"Column '{column_name}' lookup requires a library")
                if library_id not in library_ids:
                    raise ValueError(f"Column '{column_name}' references unknown library '{library_id}'")
                lookup["libraryId"] = library_id
                lookup.setdefault("caseSensitive", False)
            range_rule = validation.get("range") if isinstance(validation, dict) else None
            if range_rule is not None:
                if not isinstance(range_rule, dict):
                    raise ValueError(f"Column '{column_name}' range validation must be an object")
                range_rule.setdefault("strict", False)

    config.setdefault("version", "1.0")
    return config


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("openpyxl is required to read XLSX drillhole tables") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []
    headers = ["" if value is None else str(value) for value in rows[0]]
    out: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        item: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = raw_row[index] if index < len(raw_row) else None
            item[header] = "" if value is None else value
        if any(value not in {"", None} for value in item.values()):
            out.append(item)
    return out


def _normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        new_row: dict[str, Any] = {}
        for key, value in row.items():
            if key is None:
                continue
            new_key = _normalize_column_name(str(key))
            if new_key in {"HOLEID", "HOLE_ID", "SITEID", "BH_ID", "BHID", "BOREHOLE_ID", "HOLE"}:
                new_key = "SITE_ID"
            elif new_key in {"FROM", "START", "DEPTH_START"}:
                new_key = "DEPTH_FROM"
            elif new_key in {"TO", "END", "DEPTH_END"}:
                new_key = "DEPTH_TO"
            elif new_key in {"ID", "RECORD_ID"}:
                new_key = "ROW_ID"
            new_row[new_key] = value
        new_row["id"] = str(new_row.get("ROW_ID") or f"row-{index}")
        normalized.append(new_row)
    return normalized


def _normalize_column_name(name: str) -> str:
    normalized = name.upper().strip()
    normalized = re.sub(r"[\s().]+", "_", normalized)
    normalized = re.sub(r"_+$", "", normalized)
    return normalized


def _normalize_configs(configs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    normalized: dict[str, dict[str, Any]] = {}
    for config in configs:
        table_type = _table_type(str(config.get("tableType", "")))
        columns = []
        for col in config.get("columns", []):
            column = dict(col)
            column["columnName"] = _normalize_column_name(str(column.get("columnName", "")))
            column.setdefault("isSchemaRequired", bool(column.get("isMandatory", False)))
            column.setdefault("isMandatory", False)
            columns.append(column)
        normalized[table_type] = {**config, "tableType": table_type, "columns": columns}
    return normalized


def _table_type(name: str) -> str:
    return name.strip().replace("-", "_").upper()


def _safe_float(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _safe_site_id(row: dict[str, Any]) -> str:
    return str(row.get("SITE_ID") or row.get("HOLE_ID") or row.get("HOLEID") or row.get("id") or "Unknown")


def _validate_structure(rows: list[dict[str, Any]], config: dict[str, Any]) -> list[ValidationError]:
    if not rows:
        return []
    available = set(rows[0].keys())
    errors: list[ValidationError] = []
    for col in config.get("columns", []):
        column = str(col.get("columnName", ""))
        if col.get("isSchemaRequired") and column not in available:
            errors.append(
                ValidationError(
                    id=f"struct-{config['tableType']}-{column}",
                    table=config["tableType"],
                    row_id="HEADER",
                    site_id="SYSTEM",
                    column=column,
                    message=f"Missing Column Header: Required column '{column}' was not found in the file.",
                    severity="CRITICAL",
                    type="STRUCTURE",
                )
            )
    return errors


def _validate_values(
    rows: list[dict[str, Any]],
    config: dict[str, Any],
    libraries: list[dict[str, Any]],
) -> list[ValidationError]:
    errors: list[ValidationError] = []
    library_map = {library.get("id"): library for library in libraries}

    for row in rows:
        site_id = _safe_site_id(row)
        for col in config.get("columns", []):
            column = str(col.get("columnName", ""))
            value = row.get(column)

            if col.get("isMandatory") and _is_empty(value):
                errors.append(
                    ValidationError(
                        id=f"req-{config['tableType']}-{row['id']}-{column}",
                        table=config["tableType"],
                        row_id=row["id"],
                        site_id=site_id,
                        column=column,
                        message=f"Missing Value: Data in '{column}' cannot be empty.",
                        severity="CRITICAL",
                        type="VALUE",
                    )
                )
                continue

            if _is_empty(value):
                continue

            validation = col.get("validation") or {}
            range_rule = validation.get("range")
            if range_rule and col.get("type") != "string":
                errors.extend(_validate_range(config["tableType"], row, site_id, column, value, range_rule))

            lookup_rule = validation.get("lookup")
            if lookup_rule and col.get("type") == "string":
                library = library_map.get(lookup_rule.get("libraryId"))
                if library:
                    errors.extend(_validate_lookup(config["tableType"], row, site_id, column, value, lookup_rule, library))
    return errors


def _validate_range(
    table_type: str,
    row: dict[str, Any],
    site_id: str,
    column: str,
    value: Any,
    rule: dict[str, Any],
) -> list[ValidationError]:
    num_val = _safe_float(value)
    errors: list[ValidationError] = []
    severity = "CRITICAL" if rule.get("strict") else "WARNING"
    if "min" in rule and num_val < float(rule["min"]):
        errors.append(
            ValidationError(
                id=f"min-{table_type}-{row['id']}-{column}",
                table=table_type,
                row_id=row["id"],
                site_id=site_id,
                column=column,
                message=f"Value Too Low: {num_val:g} is below minimum {rule['min']}.",
                severity=severity,
                type="VALUE",
            )
        )
    if "max" in rule and num_val > float(rule["max"]):
        errors.append(
            ValidationError(
                id=f"max-{table_type}-{row['id']}-{column}",
                table=table_type,
                row_id=row["id"],
                site_id=site_id,
                column=column,
                message=f"Value Too High: {num_val:g} is above maximum {rule['max']}.",
                severity=severity,
                type="VALUE",
            )
        )
    return errors


def _validate_lookup(
    table_type: str,
    row: dict[str, Any],
    site_id: str,
    column: str,
    value: Any,
    rule: dict[str, Any],
    library: dict[str, Any],
) -> list[ValidationError]:
    case_sensitive = bool(rule.get("caseSensitive"))
    valid_codes = {
        str(item.get("code", "")) if case_sensitive else str(item.get("code", "")).upper()
        for item in library.get("items", [])
    }
    check_value = str(value) if case_sensitive else str(value).upper()
    if check_value in valid_codes:
        return []
    return [
        ValidationError(
            id=f"lookup-{table_type}-{row['id']}-{column}",
            table=table_type,
            row_id=row["id"],
            site_id=site_id,
            column=column,
            message=f"Invalid Code: '{value}' not found in library '{library.get('name', rule.get('libraryId'))}'.",
            severity="CRITICAL",
            type="VALUE",
        )
    ]


def _validate_collar_depths(collars: list[dict[str, Any]]) -> list[ValidationError]:
    errors = []
    for collar in collars:
        depth = _safe_float(collar.get("END_DEPTH"))
        if depth <= 0:
            errors.append(
                ValidationError(
                    id=f"collar-depth-{collar['id']}",
                    table="COLLAR",
                    row_id=collar["id"],
                    site_id=_safe_site_id(collar),
                    column="END_DEPTH",
                    message=f"Invalid Collar depth ({collar.get('END_DEPTH')}). Must be greater than zero.",
                    severity="CRITICAL",
                    type="VALUE",
                )
            )
    return errors


def _validate_integrity(
    collars: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    table_type: str,
) -> list[ValidationError]:
    valid_site_ids = {str(collar.get("SITE_ID")) for collar in collars if collar.get("SITE_ID") is not None}
    errors = []
    for row in rows:
        site_id = _safe_site_id(row)
        if site_id not in valid_site_ids:
            errors.append(
                ValidationError(
                    id=f"int-{table_type}-{row['id']}",
                    table=table_type,
                    row_id=row["id"],
                    site_id=site_id,
                    column="SITE_ID",
                    message=f"Orphan Record: Site ID '{site_id}' does not exist in Collar table.",
                    severity="CRITICAL",
                    type="INTEGRITY",
                )
            )
    return errors


def _validate_survey_eoh(collars: list[dict[str, Any]], survey: list[dict[str, Any]]) -> list[ValidationError]:
    collar_map = _collar_depth_map(collars)
    errors = []
    for row in survey:
        max_depth = collar_map.get(_safe_site_id(row))
        depth = _safe_float(row.get("DEPTH"))
        if max_depth is not None and depth > max_depth:
            errors.append(
                ValidationError(
                    id=f"eoh-surv-{row['id']}",
                    table="SURVEY",
                    row_id=row["id"],
                    site_id=_safe_site_id(row),
                    column="DEPTH",
                    message=f"Survey Depth {row.get('DEPTH')} exceeds EOH {max_depth:g}.",
                    severity="CRITICAL",
                    type="LOGIC",
                )
            )
    return errors


def _validate_eoh(
    collars: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    table_type: str,
) -> list[ValidationError]:
    collar_map = _collar_depth_map(collars)
    max_to_by_site: dict[str, float] = {}
    over_sites: set[str] = set()
    errors: list[ValidationError] = []
    tolerance = 0.01

    for row in rows:
        site_id = _safe_site_id(row)
        to_value = _safe_float(row.get("DEPTH_TO"))
        max_to_by_site[site_id] = max(to_value, max_to_by_site.get(site_id, to_value))
        max_depth = collar_map.get(site_id)
        if max_depth is not None and max_depth > 0 and to_value > max_depth + tolerance:
            over_sites.add(site_id)
            errors.append(
                ValidationError(
                    id=f"eoh-{table_type}-{row['id']}",
                    table=table_type,
                    row_id=row["id"],
                    site_id=site_id,
                    column="DEPTH_TO",
                    message=f"Depth Exceeded: 'DEPTH_TO' ({row.get('DEPTH_TO')}) exceeds Collar END_DEPTH ({max_depth:g}).",
                    severity="CRITICAL",
                    type="LOGIC",
                )
            )

    for site_id, deepest in max_to_by_site.items():
        total = collar_map.get(site_id)
        if total is not None and deepest < total - tolerance:
            errors.append(
                ValidationError(
                    id=f"eohbot-{table_type}-{site_id}",
                    table=table_type,
                    row_id="",
                    site_id=site_id,
                    column="DEPTH_TO",
                    message=f"Bottom coverage: deepest sample ({deepest:g}) is shallower than Collar END_DEPTH ({total:g}).",
                    severity="WARNING" if site_id in over_sites else "CRITICAL",
                    type="LOGIC",
                )
            )
    return errors


def _validate_intervals(rows: list[dict[str, Any]], table_type: str) -> list[ValidationError]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_safe_site_id(row), []).append(row)

    errors: list[ValidationError] = []
    for site_id, site_rows in grouped.items():
        site_rows.sort(key=lambda item: _safe_float(item.get("DEPTH_FROM")))
        for index, current in enumerate(site_rows):
            depth_from = _safe_float(current.get("DEPTH_FROM"))
            depth_to = _safe_float(current.get("DEPTH_TO"))
            if depth_from == depth_to:
                errors.append(
                    ValidationError(
                        id=f"zero-{table_type}-{current['id']}",
                        table=table_type,
                        row_id=current["id"],
                        site_id=site_id,
                        column="DEPTH_TO",
                        message=f"Zero Length: Interval {depth_from:g} to {depth_to:g} has no length.",
                        severity="WARNING",
                        type="INTERVAL",
                    )
                )
            if depth_from > depth_to:
                errors.append(
                    ValidationError(
                        id=f"inv-{table_type}-{current['id']}",
                        table=table_type,
                        row_id=current["id"],
                        site_id=site_id,
                        column="DEPTH_FROM",
                        message=f"Inverted Interval: DEPTH_FROM ({depth_from:g}) is greater than DEPTH_TO ({depth_to:g}).",
                        severity="CRITICAL",
                        type="INTERVAL",
                    )
                )
            if index == 0:
                continue
            previous = site_rows[index - 1]
            previous_to = _safe_float(previous.get("DEPTH_TO"))
            if depth_from < previous_to:
                errors.append(
                    ValidationError(
                        id=f"ovl-{table_type}-{current['id']}",
                        table=table_type,
                        row_id=current["id"],
                        site_id=site_id,
                        column="DEPTH_FROM",
                        message=f"Overlap: Starts at {depth_from:g} but previous ended at {previous_to:g}.",
                        severity="CRITICAL",
                        type="INTERVAL",
                    )
                )
            elif depth_from > previous_to:
                errors.append(
                    ValidationError(
                        id=f"gap-{table_type}-{current['id']}",
                        table=table_type,
                        row_id=current["id"],
                        site_id=site_id,
                        column="DEPTH_FROM",
                        message=f"Gap: Gap detected between {previous_to:g} and {depth_from:g}.",
                        severity="WARNING",
                        type="INTERVAL",
                    )
                )
    return errors


def _collar_depth_map(collars: list[dict[str, Any]]) -> dict[str, float]:
    return {_safe_site_id(collar): _safe_float(collar.get("END_DEPTH")) for collar in collars}


def _is_empty(value: Any) -> bool:
    return value is None or value == ""


def _render_markdown_report(summary: ValidationSummary) -> str:
    lines = [
        "# Drillhole Validation Report",
        "",
        f"- Critical errors: {summary.total_errors}",
        f"- Warnings: {summary.total_warnings}",
        "",
    ]
    if not summary.errors:
        lines.append("No validation issues found.")
        return "\n".join(lines) + "\n"

    lines.append("| Severity | Type | Table | Site | Column | Message |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for error in summary.errors:
        lines.append(
            "| "
            + " | ".join(
                [
                    error.severity,
                    error.type,
                    error.table,
                    error.site_id,
                    error.column or "",
                    error.message.replace("|", "\\|"),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"
